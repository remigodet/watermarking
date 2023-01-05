# imports
import analysis.confusion_matrix as confusion_matrix
import analysis.recall as recall
import analysis.precision as precision
import analysis.accuracy as accuracy
import analysis.metrics as metrics
import tensorflow as tf
import dataset
from tensorflow import keras

# /!\ add your models here !
import models.classifiers.classifier1 as classifier1
import models.classifiers.classifier_thomas as classifier_thomas
# /!\ add your model in this dict !
models = {
    # this is to train models automatically
    "classifier1": classifier1,
    "classifier_thomas": classifier_thomas,
}

# /!\ add your analysis modules here !
# /!\ add your model in this dict !
analysis = {
    # this is to use analysis automatically
    "metrics": metrics,
    "accuracy": accuracy,
    "precision": precision,
    "recall": recall,
    "confusion_matrix": confusion_matrix
}
class colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

# main functions
def main(model_params: dict, data_params: dict, analysis_params: dict = None) -> str:
    ''' This is the main function.
    It will create a model (training or loading from file), 
    watermark it, process it through some attacks 
    and then analyse its behaviour over a test set.
    Prints and returns the results.
    '''
    # model_setup
    model = None
    # model = model_setup(model_params, data_params, model)
    # process & analysis
    process(model, analysis_params, data_params)
    print(colors.OKGREEN + str(results) +colors.ENDC)

    ## save to Excel

    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    idx = 1
    for col_main_name in results.keys():
        for col_second_name in results[col_main_name]:
            sheet.cell(row = 1, column= idx).value = col_main_name
            sheet.cell(row = 2, column= idx).value = col_second_name
            for j, value in enumerate(results[col_main_name][col_second_name]):
                sheet.cell(row = 3+j, column= idx).value = value
            idx += 1
    workbook.save(filename=EXCEL_FILEPATH)



def process(model: tf.keras.Model, analysis_params: dict, data_params: dict) -> None:
    '''
    This is the processing part of main.py 
    where the model is subjected to changes (watermarking, attacks, retrain, ...)

    trigger_params dict is used to get the trigger set from triggerset.py
    data_params dict is used to access the dataset to train the model further and try to remove the WM

    /!\ don't forget to import those modules in main, like metrics.py

    analysis_params dict is used to know what analysis to conduct
    trigger_params dict is used to get the trigger set from triggerset.py
    data_params dict is used to access the dataset to train the model further and try to remove the WM

    Please refer to nomenclature.md on how to fill out the dictionaries
    '''
    # when implementing multiple analysis steps: loop over them
    # this is done by looping over analysis_params["processes"]
    # the dataset may be taken from the function argument or from the analysis_params dict
    # or just use classifiers module.train which will accept a model already trained ...
    for step,step_label, step_args in analysis_params:
        # colored printing
        print(colors.OKBLUE + step +": "+ colors.OKCYAN+step_label+ colors.ENDC)
        # train
        if "train" == step:
            model_params, data_params1 = step_args
            model_params = set_default_model_params(model_params)
            if data_params1 != None:
                data_params = data_params1
            model = models[model_params["classifier"]].get_model(
                model_params, data_params, model)
        # watermark
        elif "wm" == step:
            model_params, trigger_params, data_params1 = step_args
            model_params = set_default_model_params(model_params)
            if data_params1 != None:
                data_params = data_params1
            if trigger_params != None:
                model_params["wm"] = trigger_params
            model = models[model_params["classifier"]].get_model(
                model_params, data_params, model)

        #analysis
        elif step in ["confusion_matrix"]:
            data_params1, use_trigger = step_args
            analysis[step].metric(model, data_params, use_trigger)

        elif step in ["accuracy", "precision", "recall"]:
            label = step_label
            data_params1, use_trigger = step_args
            res = analysis[step].metric(model, data_params, use_trigger)

            if step in results.keys():
                if label in results[step].keys():
                    results[step][label] = results[step][label] + [res]
                elif label is not None:
                    results[step][label] = [res]
            print(colors.OKGREEN+ str(res) + colors.ENDC)

        else:
            if step_args != None:
                raise NotImplementedError(
                "analysis module behavior not defined in main.py/process")

    # for process, p_args in analysis_params["processes"]:
    #     print(process)
    #     # train
    #     if process == "train":
    #         model_params, data_params1 = p_args
    #         if data_params1 != None:
    #             data_params = data_params1
    #         model = models[model_params["classifier"]].get_model(
    #             model_params, data_params, model)
    #     # watermark
    #     if process == "wm":
    #         model_params, trigger_params, data_params1 = p_args
    #         if data_params1 != None:
    #             data_params = data_params1
    #         if trigger_params != None:
    #             model_params["wm"] = trigger_params
    #         model = models[model_params["classifier"]].get_model(
    #             model_params, data_params, model)
    # for module, a_args in analysis_params["analysis"]:
    #     print(module)
    #     try:
    #         analysis[module]
    #     except:
    #         raise Exception("no module found")
    #     if module in ["metrics"]:
    #         analysis[module].metric(model, analysis_params)
    #     elif module in ["accuracy", "precision", "recall", "confusion_matrix"]:
    #         data_params1, use_trigger = a_args
    #         print(analysis[module].metric(model, data_params, use_trigger))
    #     else:
    #         raise NotImplementedError(
    #             "analysis module behavior not defined in result")

def set_default_model_params(model_params: dict) -> dict:
    '''
    Sets missing defaults parameters a model_params dict
    '''
    if "do not train" not in model_params.keys():
        model_params["do not train"] = False

    if "carry-on" not in model_params.keys():
        model_params["carry-on"] = True
    return model_params

# to copy for new function


def func(param: type) -> None:
    ''' docstring '''
    raise NotImplementedError()


 # main
if __name__ == "__main__":

    # results

    results = {
        "accuracy": {},
    }


    # set the dict here
    hyperparams = {
        "train_ratio": 0.5,
        "val_ration": 0.3,
        "test_ration": 0.2,
        "batch_size": 32,
        'nb_epochs': 2,
        'learning_rate': 3*1e-3,
        'archi': 'boost',  # or 'dense'
        'kernel_size': (3, 3),
        'activation': 'relu',
        'nb_targets': 10,
        'nb_layers': 3,
        'add_pooling': True,
        'pooling_size': (2, 2),
        'nb_units': [32, 64, 128, 1024],
        'optimizer': keras.optimizers.Adam,
        'loss': 'sparse_categorical_crossentropy',  # 'metrics' : ['accuracy'],
    }

    data_params = {
        "dataset": "cifar-10",
        "set": "train",
        "n": 40000,
        "seed": 42
    }
    data_params_test = {
        "dataset": "cifar-10",
        "set": "test",
        "n": 3000,
        "seed": 42
    }

    trigger_params1 = {
        "n": 50,
        "nb_app_epoch": 100,
        "variance": 5,
        "from": 'dataset',
        "noise": True,
        "seed": 2
    }
    trigger_params2 = {
        "n": 50,
        "nb_app_epoch": 100,
        "variance": 5,
        "from": 'dataset',
        "noise": False,
        "seed": 3
    }

    model_params_load = {
        "saved": "WM1bis",
        "to save": None,
        "classifier": "classifier_thomas",
        "hyperparams": None,
        "wm": None,
    }

    model_params_create = {
        "saved": None,
        "to save": None,
        "classifier": "classifier_thomas",
        "hyperparams": hyperparams,
        "wm": None,
        "do not train": True,
        "carry-on": False
    }

    model_params = {
        "saved": None,
        "to save": None,
        "classifier": "classifier_thomas",
        "hyperparams": hyperparams,
        "wm": None,
    }
    
    model_params_wm1 = {
        "saved": None,
        "to save": "WM1",
        "classifier": "classifier_thomas",
        "hyperparams": hyperparams,
        "wm": trigger_params1,
    }
    model_params_wm2 = {
        "saved": None,
        "to save": "WM2",
        "classifier": "classifier_thomas",
        "hyperparams": hyperparams,
        "wm": trigger_params2,
    }

    model_params_visu = {
        "saved": "to load",
        "to save": None,
        "classifier": "classifier_thomas",
        "hyperparams": None,
        "wm": None,
    }
    analysis_params = [
            # ("train","Loading model",(model_params_load,data_params)), # load model
            # ("accuracy", "control", (data_params_test, False)),

            ("train","New model",(model_params_create,data_params)), # create new blank model
            ("accuracy", "is this model random ?", (data_params_test, False)), #to check the model is random
            

            ("train", "first training", (model_params, data_params)), #change your parameters in model_params above !
            ("accuracy", "first training", (data_params_test, False)),

            ("wm", "first  watermark training", (model_params_wm1, None, data_params)), # change your parameters in model_params above !
            ("accuracy", "first watermark training (accuracy on dataset)", (data_params_test, False)),
            ("accuracy", "first watermark training (accuracy on triggerset)", (data_params_test, trigger_params1)),
            ("confusion_matrix","confusion matrix on dataset", (data_params_test, False)),           
            ("confusion_matrix", "confusion matrix on triggerset", (data_params_test, trigger_params1)),  

            # for any precisions, check out nomenclature !

    ]

    EXCEL_FILEPATH = "results.xlsx"
    main(model_params=model_params,
         data_params=data_params,
         analysis_params=analysis_params)
    # in metrics : categories if cifar-100 ???
    print("all done")
